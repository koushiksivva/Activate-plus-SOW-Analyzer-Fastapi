import os
import io
import re
import tempfile
import pandas as pd
from concurrent.futures import ThreadPoolExecutor
import logging
import hashlib
import time
from datetime import datetime
import json

import fitz  # PyMuPDF
from fuzzywuzzy import fuzz
from task_batches import task_batches 
from dotenv import load_dotenv

# Langchain + Azure OpenAI
from langchain_openai import AzureChatOpenAI, AzureOpenAIEmbeddings
from langchain_core.prompts import PromptTemplate
from langchain.schema import HumanMessage
from langchain.text_splitter import RecursiveCharacterTextSplitter

from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import base64
from openpyxl.styles import Font
from pymongo import MongoClient
import tiktoken
from functools import lru_cache

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('app.log')
    ]
)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

AZURE_OPENAI_API_KEY = os.getenv("AZURE_OPENAI_API_KEY")
AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
AZURE_OPENAI_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION")
AZURE_OPENAI_CHAT_DEPLOYMENT = os.getenv("AZURE_OPENAI_CHAT_DEPLOYMENT")
AZURE_OPENAI_EMBEDDING_DEPLOYMENT = os.getenv("AZURE_OPENAI_EMBEDDING_DEPLOYMENT")
COSMOS_URI = os.getenv("COSMOS_URI")
COSMOS_DB = os.getenv("COSMOS_DB")
COSMOS_COLLECTION = os.getenv("COSMOS_COLLECTION")

MAX_INPUT_TOKENS = 200000
TPM_LIMIT = 245000
TPM_THRESHOLD = 0.8

# Global token usage tracker
token_stats = {
    "llm_input_tokens": 0,
    "llm_output_tokens": 0,
    "embedding_tokens": 0,
    "llm_calls": 0,
    "embedding_calls": 0,
    "start_time": time.time()
}

CHUNK_TYPES = {"TEXT": "text", "IMAGE": "image"}

# Initialize Azure OpenAI
try:
    llm = AzureChatOpenAI(
        openai_api_version=AZURE_OPENAI_API_VERSION,
        azure_deployment=AZURE_OPENAI_CHAT_DEPLOYMENT,
        azure_endpoint=AZURE_OPENAI_ENDPOINT,
        api_key=AZURE_OPENAI_API_KEY,
        temperature=0.1
    )
    vision_llm = AzureChatOpenAI(
        openai_api_version=AZURE_OPENAI_API_VERSION,
        azure_deployment="gpt-4o",
        azure_endpoint=AZURE_OPENAI_ENDPOINT,
        api_key=AZURE_OPENAI_API_KEY,
        temperature=0.1
    )
except Exception as e:
    logger.error(f"Failed to initialize Azure OpenAI: {str(e)}")
    raise

try:
    embedding_model = AzureOpenAIEmbeddings(
        azure_deployment=AZURE_OPENAI_EMBEDDING_DEPLOYMENT,
        openai_api_version=AZURE_OPENAI_API_VERSION,
        azure_endpoint=AZURE_OPENAI_ENDPOINT,
        api_key=AZURE_OPENAI_API_KEY,
        model=AZURE_OPENAI_EMBEDDING_DEPLOYMENT
    )
    test_embedding = embedding_model.embed_query("test")
    logger.info("Azure OpenAI Embeddings connection test successful")
except Exception as e:
    logger.error(f"Failed to initialize Azure OpenAI Embeddings: {str(e)}")
    raise

try:
    client = MongoClient(COSMOS_URI)
    db = client[COSMOS_DB]
    collection = db[COSMOS_COLLECTION]
    collection.count_documents({})
    logger.info("Cosmos DB connection successful")
except Exception as e:
    logger.error(f"Failed to connect to Cosmos DB: {str(e)}")
    raise

@lru_cache(maxsize=10000)
def get_query_embedding(query: str):
    return embedding_model.embed_query(query)

@lru_cache(maxsize=20000)
def cached_doc_embedding(text: str):
    return embedding_model.embed_documents([text])[0]

def check_tpm_limit():
    current_time = time.time()
    elapsed_minutes = (current_time - token_stats["start_time"]) / 60
    if elapsed_minutes > 0:
        current_tpm = (token_stats["llm_input_tokens"] + token_stats["llm_output_tokens"]) / elapsed_minutes
        if current_tpm > (TPM_LIMIT * TPM_THRESHOLD):
            sleep_time = 60 - (elapsed_minutes % 60)
            logger.info(f"Approaching TPM limit ({current_tpm:.0f}). Pausing for {sleep_time:.1f}s")
            time.sleep(sleep_time)
            token_stats["start_time"] = time.time()

def count_tokens(text: str, model: str = "gpt-4o"):
    if not text or not isinstance(text, str):
        return 0
    try:
        encoding = tiktoken.encoding_for_model(model)
    except KeyError:
        encoding = tiktoken.get_encoding("cl100k_base")
    return len(encoding.encode(text))

def truncate_context(context, max_tokens=MAX_INPUT_TOKENS):
    if not context or not isinstance(context, str):
        return ""
    encoding = tiktoken.get_encoding("cl100k_base")
    tokens = encoding.encode(context)
    if len(tokens) <= max_tokens:
        return context
    keep_start = int(max_tokens * 0.7)
    keep_end = max_tokens - keep_start
    truncated_tokens = tokens[:keep_start] + tokens[-keep_end:]
    return encoding.decode(truncated_tokens)

def create_vector_index():
    try:
        existing_indexes = list(collection.list_indexes())
        vector_index_exists = any(idx.get('name') == 'vectorSearchIndex' for idx in existing_indexes)
        if not vector_index_exists:
            index_definition = {
                "createIndexes": COSMOS_COLLECTION,
                "indexes": [
                    {
                        "name": "vectorSearchIndex",
                        "key": {"embedding": "cosmosSearch"},
                        "cosmosSearchOptions": {
                            "kind": "vector-ivf",
                            "numLists": 1,
                            "similarity": "COS",
                            "dimensions": 1536
                        }
                    }
                ]
            }
            db.command(index_definition)
        collection.create_index([("text_hash", 1)], name="text_hash_index", unique=False, sparse=True)
        collection.create_index([("desc_hash", 1)], name="desc_hash_index", unique=False, sparse=True)
        collection.create_index([("document_id", 1)], name="document_id_index")
        return True
    except Exception as e:
        logger.error(f"Index creation failed: {e}")
        return False

vector_index_available = create_vector_index()

def extract_pdf_content_pymupdf(pdf_path):
    text_content = ""
    images_content = []
    try:
        doc = fitz.open(pdf_path)
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            text = page.get_text()
            if text and text.strip():
                text_content += f"\n\n--- Page {page_num + 1} ---\n\n{text}"
            try:
                tables = page.find_tables()
                for table_idx, table in enumerate(tables):
                    table_data = table.extract()
                    if table_data and len(table_data) > 1:
                        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(table_data[0])]
                        df = pd.DataFrame(table_data[1:], columns=headers)
                        text_content += f"\n\n--- Page {page_num + 1} Table {table_idx + 1} ---\n\n{df.to_string(index=False)}"
            except Exception as table_error:
                logger.warning(f"Error extracting tables from page {page_num + 1}: {table_error}")
            images = page.get_images(full=True)
            for img_index, img in enumerate(images):
                try:
                    xref = img[0]
                    base_image = doc.extract_image(xref)
                    image_bytes = base_image["image"]
                    image_b64 = base64.b64encode(image_bytes).decode('utf-8')
                    vision_description = analyze_image_for_durations(image_b64)
                    image_info = {
                        "page": page_num + 1,
                        "index": img_index + 1,
                        "data": image_b64,
                        "description": vision_description or f"Page {page_num + 1} - Image {img_index + 1}"
                    }
                    images_content.append(image_info)
                    if vision_description and any(keyword in vision_description.lower() for keyword in ['weeks', 'months', 'phase', 'sprint']):
                        text_content += f"\n\n--- Page {page_num + 1} Image Analysis ---\n\n{vision_description}"
                except Exception as img_error:
                    logger.warning(f"Error processing image {img_index + 1} on page {page_num + 1}: {img_error}")
        doc.close()
        return text_content, images_content
    except Exception as e:
        logger.error(f"Error extracting PDF content with PyMuPDF: {e}")
        return "", []

def analyze_image_for_durations(image_b64):
    try:
        check_tpm_limit()
        prompt = """
        Analyze this image for project timeline or Gantt chart information. Look for:
        1. Phase durations (PREP, EXPLORE, REALIZE, DEPLOY, RUN phases)
        2. Sprint durations or counts
        3. Timeline bars showing months/weeks
        4. Any duration numbers or time spans
        If this appears to be a timeline/Gantt chart:
        - Count bar lengths or time spans
        - Sum sprint durations (e.g., 7 sprints × 3 weeks = 21 weeks)
        - Convert to months if needed (4 weeks ≈ 1 month)
        Output format: "Phase: Duration" for each phase found, or "No timeline data" if none found.
        Be concise and focus only on duration information.
        """
        input_tokens = count_tokens(prompt, model="gpt-4o")
        token_stats["llm_input_tokens"] += input_tokens
        token_stats["llm_calls"] += 1
        messages = [
            HumanMessage(
                content=[
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_b64}"}}
                ]
            )
        ]
        response = vision_llm(messages)
        output_tokens = count_tokens(response.content, model="gpt-4o")
        token_stats["llm_output_tokens"] += output_tokens
        return response.content
    except Exception as e:
        logger.warning(f"Vision analysis failed: {e}")
        return None

def generate_document_id(pdf_content):
    if not pdf_content:
        pdf_content = ""
    if not isinstance(pdf_content, str):
        pdf_content = str(pdf_content)
    normalized = normalize_and_clean_text(pdf_content)
    return hashlib.md5(normalized.encode('utf-8')).hexdigest()

def normalize_and_clean_text(text):
    if not text:
        return ""
    text = str(text).lower().strip()
    text = re.sub(r'[^\w\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text

def check_existing_chunks(document_id):
    try:
        return collection.count_documents({"document_id": document_id}) > 0
    except Exception as e:
        logger.error(f"Error checking existing chunks: {e}")
        return False

def get_existing_embedding(text):
    try:
        text_hash = hashlib.md5(normalize_and_clean_text(text).encode('utf-8')).hexdigest()
        existing = collection.find_one({
            "text_hash": text_hash,
            "chunk_type": CHUNK_TYPES["TEXT"]
        })
        return existing["embedding"] if existing and "embedding" in existing else None
    except Exception as e:
        logger.error(f"Error getting existing embedding: {e}")
        return None

def store_chunks_in_cosmos(text_chunks, image_chunks, document_id):
    try:
        existing_chunks = {
            doc["chunk_id"]: doc for doc in collection.find(
                {"document_id": document_id},
                {"chunk_id": 1, "embedding": 1}
            )
        }
        documents_to_insert = []
        for i, chunk in enumerate(text_chunks):
            if not chunk.strip():
                continue
            chunk_id = f"{document_id}_text_{i}"
            if chunk_id in existing_chunks:
                logger.info(f"Skipping existing text chunk {i}")
                continue
            text_hash = hashlib.md5(normalize_and_clean_text(chunk).encode('utf-8')).hexdigest()
            embedding = get_existing_embedding(chunk) or cached_doc_embedding(chunk)
            documents_to_insert.append({
                "document_id": document_id,
                "chunk_id": chunk_id,
                "text": chunk,
                "text_hash": text_hash,
                "embedding": embedding,
                "chunk_index": i,
                "chunk_type": CHUNK_TYPES["TEXT"],
                "created_at": pd.Timestamp.now().isoformat()
            })
        for i, image_info in enumerate(image_chunks):
            if not image_info.get("description"):
                continue
            chunk_id = f"{document_id}_image_{i}"
            if chunk_id in existing_chunks:
                logger.info(f"Skipping existing image chunk {i}")
                continue
            description = image_info["description"]
            desc_hash = hashlib.md5(normalize_and_clean_text(description).encode('utf-8')).hexdigest()
            embedding = get_existing_embedding(description) or cached_doc_embedding(description)
            documents_to_insert.append({
                "document_id": document_id,
                "chunk_id": chunk_id,
                "image_data": image_info["data"],
                "image_description": description,
                "desc_hash": desc_hash,
                "embedding": embedding,
                "chunk_index": i,
                "chunk_type": CHUNK_TYPES["IMAGE"],
                "page": image_info["page"],
                "created_at": pd.Timestamp.now().isoformat()
            })
        if documents_to_insert:
            collection.insert_many(documents_to_insert, ordered=False)
            logger.info(f"Inserted {len(documents_to_insert)} new chunks for {document_id}")
        else:
            logger.info(f"No new chunks to insert for {document_id}")
        return True
    except Exception as e:
        logger.error(f"Error storing chunks in Cosmos DB: {e}")
        return False

def similarity_search_cosmos(query_text, document_id, k=5):
    try:
        results = []
        if vector_index_available:
            try:
                query_embedding = get_query_embedding(query_text)
                pipeline = [
                    {
                        "$search": {
                            "cosmosSearch": {
                                "vector": query_embedding,
                                "path": "embedding",
                                "k": k
                            }
                        }
                    },
                    {"$match": {"document_id": document_id}},
                    {"$project": {"text": 1, "chunk_type": 1, "image_description": 1, "image_data": 1, "_id": 0}}
                ]
                vector_results = list(collection.aggregate(pipeline))
                for doc in vector_results:
                    if doc.get("chunk_type") == CHUNK_TYPES["TEXT"]:
                        results.append({"page_content": doc["text"], "chunk_type": CHUNK_TYPES["TEXT"]})
                    elif doc.get("chunk_type") == CHUNK_TYPES["IMAGE"]:
                        results.append({
                            "page_content": doc["image_description"],
                            "chunk_type": CHUNK_TYPES["IMAGE"],
                            "image_data": doc.get("image_data", "")
                        })
                if results:
                    return results
            except Exception as ve:
                logger.warning(f"Vector search failed: {ve}")
        query_words = query_text.lower().split()[:5]
        text_search_query = {
            "document_id": document_id,
            "$or": [{"text": {"$regex": word, "$options": "i"}} for word in query_words]
        }
        text_results = list(collection.find(
            text_search_query,
            {"text": 1, "chunk_type": 1, "image_description": 1, "image_data": 1, "_id": 0}
        ).limit(k))
        for doc in text_results:
            if doc.get("chunk_type") == CHUNK_TYPES["TEXT"]:
                results.append({"page_content": doc["text"], "chunk_type": CHUNK_TYPES["TEXT"]})
            elif doc.get("chunk_type") == CHUNK_TYPES["IMAGE"]:
                results.append({
                    "page_content": doc["image_description"],
                    "chunk_type": CHUNK_TYPES["IMAGE"],
                    "image_data": doc.get("image_data", "")
                })
        if not results:
            logger.warning("Falling back to any document chunks")
            fallback_docs = list(collection.find(
                {"document_id": document_id},
                {"text": 1, "chunk_type": 1, "image_description": 1, "image_data": 1, "_id": 0}
            ).limit(k))
            for doc in fallback_docs:
                if doc.get("chunk_type") == CHUNK_TYPES["TEXT"]:
                    results.append({"page_content": doc["text"], "chunk_type": CHUNK_TYPES["TEXT"]})
                elif doc.get("chunk_type") == CHUNK_TYPES["IMAGE"]:
                    results.append({
                        "page_content": doc["image_description"],
                        "chunk_type": CHUNK_TYPES["IMAGE"],
                        "image_data": doc.get("image_data", "")
                    })
        return results
    except Exception as e:
        logger.error(f"Error in Cosmos DB search: {str(e)}")
        return []

duration_template = PromptTemplate.from_template("""
[Document Content]
{context}

INSTRUCTIONS:
1. Extract durations for phases (PREP, EXPLORE, REALIZE, DEPLOY, RUN) in ANY format: text, tables, Gantt charts, timelines.
2. For Gantt charts/timelines/images:
   - Interpret bars/lines: Count months/weeks spanned (e.g., bar from Month 3 to 10 = 7 months)
   - Sum sprints/sub-phases (e.g., Sprint 1: 4 weeks + Sprints 2-7: 3 weeks each = 4 + 18 = 22 weeks = ~5 months)
   - Look for visual timeline bars showing duration spans
   - Count numbered sprints or phases and multiply by duration
   - Assume months if just numbers; convert weeks to months (4 weeks ≈ 1 month)
3. Phase name variations:
   - PREP = Prep/Preparation/Planning
   - EXPLORE = Explore/Discovery/Analysis
   - REALIZE = Realize/Build/Development/Sprints/Construction
   - DEPLOY = Deploy/UAT/Testing/Training/Go-Live
   - RUN = Run/Hypercare/Support/Maintenance
4. Sum all sub-activities within each phase
5. Return empty string if phase not found

OUTPUT ONLY:
{{
    "durations": {{
        "PREP": "X weeks",
        "EXPLORE": "X weeks",
        "REALIZE": "X weeks",
        "DEPLOY": "X weeks",
        "RUN": "X weeks"
    }}
}}
""")

task_template = PromptTemplate.from_template("""
[Document Content]
{context}

INSTRUCTIONS:
1. Carefully analyze ONLY the provided Statement of Work (SOW) content.
2. For each task below, determine if it is EXPLICITLY mentioned in the SOW.
3. Only answer "yes" if the task is clearly stated in the document.
4. Answer "no" if the task is not mentioned or only vaguely referenced.
5. Do not infer or assume anything not explicitly stated.
6. Output must be JSON with the following strict structure:
{{
    "tasks": {{
        "task1": "yes/no",
        "task2": "yes/no",
        ...
    }}
}}

RULES:
- Be extremely precise - only "yes" for exact matches
- Ignore similar-sounding but different tasks
- Do not include any explanations
- Do not add any information not in the SOW

TASKS TO CHECK:
{tasks_string}

STRICT JSON OUTPUT:
""")

def safe_invoke(prompt, max_retries=2):
    retries = 0
    while retries < max_retries:
        try:
            check_tpm_limit()
            truncated_prompt = truncate_context(prompt, MAX_INPUT_TOKENS)
            input_tokens = count_tokens(truncated_prompt, model="gpt-4o")
            token_stats["llm_input_tokens"] += input_tokens
            token_stats["llm_calls"] += 1
            messages = [HumanMessage(content=truncated_prompt)]
            response_obj = llm(messages)
            output_text = response_obj.content
            output_tokens = count_tokens(output_text, model="gpt-4o")
            token_stats["llm_output_tokens"] += output_tokens
            match = re.search(r'\{.*\}', output_text, re.DOTALL)
            if match:
                return json.loads(match.group())
        except Exception as e:
            logger.error(f"LLM Error (Attempt {retries+1}): {e}")
            retries += 1
    logger.error("Max retries reached for LLM invocation")
    return None

def verify_substring_match(task, context):
    if not task or not context:
        return "no"
    try:
        norm_task = normalize_and_clean_text(task)
        norm_context = normalize_and_clean_text(context)
        return "yes" if norm_task in norm_context else "no"
    except Exception as e:
        logger.error(f"Error in substring match: {e}")
        return "no"

def fuzzy_match_optimized(task, pdf_text, threshold=75, window=5000, step=4000):
    if not task or not pdf_text:
        return "no"
    try:
        task_str = normalize_and_clean_text(task)
        pdf_str = normalize_and_clean_text(pdf_text)
        best_score = 0
        for start in range(0, len(pdf_str), step):
            snippet = pdf_str[start:start+window]
            score = fuzz.partial_ratio(task_str, snippet)
            if score > best_score:
                best_score = score
            if best_score >= threshold:
                return "yes"
        return "yes" if best_score >= threshold else "no"
    except Exception as e:
        logger.error(f"Error in fuzzy_match_optimized: {e}")
        return "no"

def extract_durations_optimized(pdf_text):
    try:
        duration_context = truncate_context(pdf_text, 15000)
        prompt = duration_template.format(context=duration_context)
        response = safe_invoke(prompt)
        if response and "durations" in response:
            return response["durations"]
    except Exception as e:
        logger.error(f"Error extracting durations: {e}")
    return {
        "PREP": "",
        "EXPLORE": "",
        "REALIZE": "",
        "DEPLOY": "",
        "RUN": ""
    }

def process_batch_with_fallback(sub_batch, document_id, durations, normalized_pdf_text, pdf_text):
    try:
        batch_results = []
        docs = []
        substring_flags = {}
        for heading, task in sub_batch:
            if not task:
                continue
            task = str(task)
            try:
                query_embedding = get_query_embedding(task)
                pipeline = [
                    {
                        "$search": {
                            "cosmosSearch": {
                                "vector": query_embedding,
                                "path": "embedding",
                                "k": 3
                            }
                        }
                    },
                    {"$match": {"document_id": document_id}},
                    {"$project": {"text": 1, "chunk_type": 1, "image_description": 1, "image_data": 1, "_id": 0}}
                ]
                vector_results = list(collection.aggregate(pipeline))
                for doc in vector_results:
                    if doc.get("text"):
                        docs.append(doc["text"])
                    elif doc.get("image_description"):
                        docs.append(doc["image_description"])
            except Exception as search_error:
                logger.warning(f"Vector search failed for '{task}': {search_error}")
            substring_flags[task] = verify_substring_match(task, normalized_pdf_text)
        unique_docs = list(set(docs))
        combined_context = "\n".join(unique_docs[:5])
        tasks_string = "\n".join([f"task{i+1}: {str(task)}" for i, (_, task) in enumerate(sub_batch) if task])
        llm_response = None
        if combined_context and tasks_string:
            prompt = task_template.format(context=combined_context, tasks_string=tasks_string)
            llm_response = safe_invoke(prompt)
        for i, (heading, task) in enumerate(sub_batch):
            task_str = str(task) if task else ""
            final_answer = "no"
            if substring_flags.get(task_str, "no") == "yes":
                final_answer = "yes"
            elif llm_response and "tasks" in llm_response:
                task_values = list(llm_response["tasks"].values())
                if i < len(task_values) and str(task_values[i]).lower() == "yes":
                    final_answer = "yes"
            if final_answer == "no":
                final_answer = fuzzy_match_optimized(task_str, pdf_text)
            batch_results.append({
                "Heading": str(heading) if heading else "Unknown",
                "Task": task_str,
                "Present": final_answer,
                **durations
            })
        return batch_results
    except Exception as e:
        logger.error(f"Error processing batch: {e}")
        return [{"Heading": str(h), "Task": str(t), "Present": "error", **durations} for h, t in sub_batch]

def process_pdf_safely(uploaded_file):
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_pdf_file:
            tmp_pdf_file.write(uploaded_file.file.read() if hasattr(uploaded_file, "file") else uploaded_file.read())
            tmp_pdf_path = tmp_pdf_file.name
        pdf_text, images_content = extract_pdf_content_pymupdf(tmp_pdf_path)
        if not pdf_text or not pdf_text.strip():
            return None
        normalized_pdf_text = normalize_and_clean_text(pdf_text)
        if not normalized_pdf_text.strip():
            return None
        return pdf_text, normalized_pdf_text, tmp_pdf_path, images_content
    except Exception as e:
        logger.error(f"Error in process_pdf_safely: {str(e)}")
        return None

def create_excel_with_formatting(df, durations, output_file, activity_column_width=50):
    results_dict = {}
    for heading in df['Heading'].unique():
        tasks = df[df['Heading'] == heading]
        task_dict = {task: present for task, present in zip(tasks['Task'], tasks['Present'])}
        results_dict[heading] = task_dict
    yes_rows = []
    no_rows = []
    for category, tasks in results_dict.items():
        if category.upper() in durations:
            yes_rows.append({"Phase": category.upper(), "Duration": durations.get(category.upper(), ""), "Activity": ""})
        else:
            yes_rows.append({"Phase": "", "Duration": "", "Activity": category})
        for task, present in tasks.items():
            if present == "yes":
                yes_rows.append({"Phase": "", "Duration": "", "Activity": task})
        yes_rows.append({"Phase": "", "Duration": "", "Activity": ""})
        no_tasks = [task for task, present in tasks.items() if present == "no"]
        if no_tasks:
            if category.upper() in durations:
                no_rows.append({"Phase": category.upper(), "Duration": durations.get(category.upper(), ""), "Activity": ""})
            else:
                no_rows.append({"Phase": "", "Duration": "", "Activity": category})
            for task in no_tasks:
                no_rows.append({"Phase": "", "Duration": "", "Activity": task})
            no_rows.append({"Phase": "", "Duration": "", "Activity": ""})
    yes_df = pd.DataFrame(yes_rows)
    no_df = pd.DataFrame(no_rows)
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        if not yes_df.empty:
            yes_df.to_excel(writer, index=False, startrow=1, sheet_name="Yes_Tasks")
            workbook = writer.book
            worksheet = writer.sheets["Yes_Tasks"]
            activity_column_index = yes_df.columns.get_loc("Activity") + 1
            column_letter = get_column_letter(activity_column_index)
            worksheet.column_dimensions[column_letter].width = activity_column_width
            bold_font = Font(bold=True)
            for row in range(2, len(yes_rows) + 2):
                phase_cell = worksheet.cell(row=row, column=1)
                activity_cell = worksheet.cell(row=row, column=3)
                if phase_cell.value in durations.keys():
                    phase_cell.font = bold_font
                    worksheet.cell(row=row, column=2).font = bold_font
                if activity_cell.value in results_dict.keys() and activity_cell.value.upper() not in durations:
                    activity_cell.font = bold_font
            table = Table(displayName="PresentTasksTable", ref=worksheet.dimensions)
            style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            table.tableStyleInfo = style
            worksheet.add_table(table)
        if not no_df.empty:
            no_df.to_excel(writer, index=False, startrow=1, sheet_name="No_Tasks")
            worksheet_no = writer.sheets["No_Tasks"]
            activity_column_index = no_df.columns.get_loc("Activity") + 1
            column_letter = get_column_letter(activity_column_index)
            worksheet_no.column_dimensions[column_letter].width = activity_column_width
            for row in range(2, len(no_rows) + 2):
                phase_cell = worksheet_no.cell(row=row, column=1)
                activity_cell = worksheet_no.cell(row=row, column=3)
                if phase_cell.value in durations.keys():
                    phase_cell.font = bold_font
                    worksheet_no.cell(row=row, column=2).font = bold_font
                if activity_cell.value in results_dict.keys() and activity_cell.value.upper() not in durations:
                    activity_cell.font = bold_font
            table_no = Table(displayName="MissingTasksTable", ref=worksheet_no.dimensions)
            style_no = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            table_no.tableStyleInfo = style_no
            worksheet_no.add_table(table_no)

